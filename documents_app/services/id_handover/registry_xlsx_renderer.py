from __future__ import annotations

import math
from functools import lru_cache
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.formula.translate import Translator
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from PIL import ImageFont


@dataclass(slots=True)
class PlaceholderCellMatch:
    """
    Найденный плейсхолдер в workbook.

    sheet (лист Excel)
    row (номер строки)
    column (номер колонки)
    value (исходное значение ячейки)
    """
    sheet: Worksheet
    row: int
    column: int
    value: str


class RegistryXlsxRendererError(Exception):
    """Базовая ошибка рендера XLSX-реестра."""


class RegistryXlsxRendererValidationError(RegistryXlsxRendererError):
    """Ошибка структуры шаблона XLSX-реестра."""


class RegistryXlsxRenderer:
    """
    Рендерит XLSX-реестр по уже готовому context.

    Поддерживает:
    - замену шапочных плейсхолдеров
    - вставку табличной части по rows
    - сохранение исходного форматирования шаблона
    - bold для строк акта
    - перенос формулы в колонке "Страница"

    Ожидаемые специальные плейсхолдеры:
    - {{documents}}
    - {{sheets}}

    ВАЖНО:
    - оба плейсхолдера должны находиться на одной строке шаблона
    - documents/sheets — это anchors табличной части
    - значение для "Страница" не заполняем руками, а копируем формулу шаблонной строки
    - renderer НЕ строит context сам
    """

    DOCUMENTS_PLACEHOLDER = "{{documents}}"
    SHEETS_PLACEHOLDER = "{{sheets}}"

    HEADER_PLACEHOLDERS = {
        "{{tech_customer_org_short}}": "tech_customer_org_short",
        "{{builder_rep_org_short}}": "builder_rep_org_short",
        "{{other_rep_org_short}}": "other_rep_org_short",
        "{{contractor_rep_org_short}}": "contractor_rep_org_short",
        "{{project_line_full_name}}": "project_line_full_name",
        "{{project_stage_full_name}}": "project_stage_full_name",
        "{{project_full_code}}": "project_full_code",
        "{{project_plot_full_name}}": "project_plot_full_name",
        "{{project_construction}}": "project_construction",
        "{{executive_or_and_working_documentation_registry}}": "executive_or_and_working_documentation_registry",
    }

    FONT_FILE_CANDIDATES = {
        "calibri": (
            "calibri.ttf",
            "calibrib.ttf",
            "Carlito-Regular.ttf",
            "Carlito-Bold.ttf",
            "LiberationSans-Regular.ttf",
            "LiberationSans-Bold.ttf",
            "DejaVuSans.ttf",
            "DejaVuSans-Bold.ttf",
        ),
        "times new roman": (
            "times.ttf",
            "timesbd.ttf",
            "LiberationSerif-Regular.ttf",
            "LiberationSerif-Bold.ttf",
            "DejaVuSerif.ttf",
            "DejaVuSerif-Bold.ttf",
        ),
        "arial": (
            "arial.ttf",
            "arialbd.ttf",
            "LiberationSans-Regular.ttf",
            "LiberationSans-Bold.ttf",
            "DejaVuSans.ttf",
            "DejaVuSans-Bold.ttf",
        ),
        "tahoma": (
            "tahoma.ttf",
            "tahomabd.ttf",
            "DejaVuSans.ttf",
            "DejaVuSans-Bold.ttf",
        ),
        "cambria": (
            "cambria.ttc",
            "cambriab.ttf",
            "Caladea-Regular.ttf",
            "Caladea-Bold.ttf",
            "DejaVuSerif.ttf",
            "DejaVuSerif-Bold.ttf",
        ),
    }
    FONT_SEARCH_DIRS = (
        Path("C:/Windows/Fonts"),
        Path("/usr/share/fonts"),
        Path("/usr/local/share/fonts"),
    )

    def render(
        self,
        *,
        context: dict[str, Any],
        template_path: str | Path,
        output_path: str | Path,
    ) -> Path:
        """
        Публичная точка входа.

        Renderer работает только по уже собранному context.
        """
        return self.render_from_context(
            context=context,
            template_path=template_path,
            output_path=output_path,
        )

    def render_from_context(
        self,
        *,
        context: dict[str, Any],
        template_path: str | Path,
        output_path: str | Path,
    ) -> Path:
        """
        Рендерит XLSX по уже готовому context.
        """
        template_path = Path(template_path)
        output_path = Path(output_path)

        if not template_path.exists():
            raise RegistryXlsxRendererValidationError(
                f"Шаблон не найден: {template_path}"
            )

        workbook = load_workbook(template_path)
        worksheet = self._resolve_target_sheet(workbook)

        placeholders = context.get("placeholders", {})
        rows = context.get("rows", [])

        if not rows:
            raise RegistryXlsxRendererValidationError(
                "Контекст реестра не содержит строк rows."
            )

        self._replace_header_placeholders(
            worksheet=worksheet,
            placeholders=placeholders,
        )

        documents_match = self._find_placeholder_cell(
            worksheet=worksheet,
            placeholder=self.DOCUMENTS_PLACEHOLDER,
        )
        sheets_match = self._find_placeholder_cell(
            worksheet=worksheet,
            placeholder=self.SHEETS_PLACEHOLDER,
        )

        if documents_match.row != sheets_match.row:
            raise RegistryXlsxRendererValidationError(
                "Плейсхолдеры {{documents}} и {{sheets}} должны находиться на одной строке."
            )

        template_row_idx = documents_match.row
        documents_col_idx = documents_match.column
        number_col_idx = documents_col_idx - 1
        sheets_col_idx = sheets_match.column
        page_col_idx = sheets_col_idx + 1

        preallocated_rows_count = self._count_preallocated_table_rows(
            worksheet=worksheet,
            template_row_idx=template_row_idx,
            page_col_idx=page_col_idx,
        )
        additional_rows_count = max(0, len(rows) - preallocated_rows_count)

        if additional_rows_count > 0:
            worksheet.insert_rows(
                template_row_idx + preallocated_rows_count,
                amount=additional_rows_count,
            )
            self._replicate_template_row_block(
                worksheet=worksheet,
                template_row_idx=template_row_idx,
                rows_count=preallocated_rows_count + additional_rows_count,
            )

        for offset, row_context in enumerate(rows):
            target_row_idx = template_row_idx + offset
            self._render_single_registry_row(
                worksheet=worksheet,
                target_row_idx=target_row_idx,
                number_col_idx=number_col_idx,
                documents_col_idx=documents_col_idx,
                sheets_col_idx=sheets_col_idx,
                row_context=row_context,
            )

        self._apply_documents_rows_heights(
            worksheet=worksheet,
            template_row_idx=template_row_idx,
            documents_col_idx=documents_col_idx,
            sheets_col_idx=sheets_col_idx,
            rows_count=len(rows),
        )

        unused_preallocated_rows_count = max(0, preallocated_rows_count - len(rows))
        if unused_preallocated_rows_count > 0:
            self._delete_rows_preserving_footer_layout(
                worksheet=worksheet,
                delete_start_row=template_row_idx + len(rows),
                amount=unused_preallocated_rows_count,
            )

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)

        return output_path

    def _resolve_target_sheet(self, workbook):
        """
        Пока используем active sheet.

        Если позже в шаблоне появится отдельное имя листа,
        сюда можно добавить явный выбор по имени.
        """
        return workbook.active

    def _replace_header_placeholders(
        self,
        *,
        worksheet: Worksheet,
        placeholders: dict[str, Any],
    ) -> None:
        """
        Заменяет шапочные плейсхолдеры по всему листу.

        Стиль ячеек не меняется — меняется только value.
        """
        placeholder_value_map = {
            placeholder: str(placeholders.get(context_key, "") or "")
            for placeholder, context_key in self.HEADER_PLACEHOLDERS.items()
        }

        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if not isinstance(cell.value, str):
                    continue

                new_value = cell.value
                changed = False

                for placeholder, replacement in placeholder_value_map.items():
                    if placeholder in new_value:
                        new_value = new_value.replace(placeholder, replacement)
                        changed = True

                if changed:
                    cell.value = new_value

    def _find_placeholder_cell(
        self,
        *,
        worksheet: Worksheet,
        placeholder: str,
    ) -> PlaceholderCellMatch:
        """
        Ищет ячейку, которая содержит точный плейсхолдер.
        """
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                if cell.value == placeholder:
                    return PlaceholderCellMatch(
                        sheet=worksheet,
                        row=cell.row,
                        column=cell.column,
                        value=cell.value,
                    )

        raise RegistryXlsxRendererValidationError(
            f"В шаблоне не найден плейсхолдер {placeholder}"
        )

    def _replicate_template_row_block(
        self,
        *,
        worksheet: Worksheet,
        template_row_idx: int,
        rows_count: int,
    ) -> None:
        """
        Копирует стиль/формулы/мерджи шаблонной строки на вставленные строки.

        Логика:
        - template_row_idx уже существует
        - ниже уже вставлены новые пустые строки
        - нужно протянуть на них:
          - стили ячеек
          - высоту строки
          - формулы/значения шаблонной строки
          - merged ranges вида B:D на шаблонной строке
        """
        max_col = worksheet.max_column
        template_row_height = worksheet.row_dimensions[template_row_idx].height

        merged_ranges_for_template_row = self._get_single_row_merged_ranges(
            worksheet=worksheet,
            row_idx=template_row_idx,
        )

        for target_row_idx in range(template_row_idx + 1, template_row_idx + rows_count):
            if template_row_height is not None:
                worksheet.row_dimensions[target_row_idx].height = template_row_height

            for col_idx in range(1, max_col + 1):
                source_cell = worksheet.cell(row=template_row_idx, column=col_idx)
                target_cell = worksheet.cell(row=target_row_idx, column=col_idx)

                if isinstance(source_cell, MergedCell):
                    continue

                self._copy_cell_style(source_cell=source_cell, target_cell=target_cell)
                self._copy_cell_value_with_formula_translation(
                    source_cell=source_cell,
                    target_cell=target_cell,
                )

            for merged_range in merged_ranges_for_template_row:
                row_offset = target_row_idx - template_row_idx
                new_min_row = merged_range.min_row + row_offset
                new_max_row = merged_range.max_row + row_offset
                new_range = (
                    f"{worksheet.cell(new_min_row, merged_range.min_col).coordinate}:"
                    f"{worksheet.cell(new_max_row, merged_range.max_col).coordinate}"
                )
                worksheet.merge_cells(new_range)

    def _get_single_row_merged_ranges(
        self,
        *,
        worksheet: Worksheet,
        row_idx: int,
    ):
        """
        Возвращает merged ranges, которые полностью лежат в одной template-строке.

        Для нашего шаблона это нужно, чтобы копировать объединение B:D на каждую новую строку.
        """
        result = []
        for merged_range in worksheet.merged_cells.ranges:
            if merged_range.min_row == row_idx and merged_range.max_row == row_idx:
                result.append(merged_range)
        return result

    def _copy_cell_style(self, *, source_cell, target_cell) -> None:
        """
        Копирует стиль ячейки без изменения значения.
        """
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.protection:
            target_cell.protection = copy(source_cell.protection)

    def _copy_cell_value_with_formula_translation(self, *, source_cell, target_cell) -> None:
        """
        Копирует значение шаблонной ячейки.

        Если в source формула, пытаемся сдвинуть относительные ссылки на новую строку.
        Это особенно важно для колонки "Страница".
        """
        source_value = source_cell.value

        if not isinstance(source_value, str) or not source_value.startswith("="):
            target_cell.value = source_value
            return

        try:
            translated_formula = Translator(
                formula=source_value,
                origin=source_cell.coordinate,
            ).translate_formula(target_cell.coordinate)
            target_cell.value = translated_formula
        except Exception:
            target_cell.value = source_value

    def _render_single_registry_row(
        self,
        *,
        worksheet: Worksheet,
        target_row_idx: int,
        number_col_idx: int,
        documents_col_idx: int,
        sheets_col_idx: int,
        row_context: dict[str, Any],
    ) -> None:
        """
        Заполняет одну строку реестра.
        """
        number_cell = worksheet.cell(row=target_row_idx, column=number_col_idx)
        documents_cell = worksheet.cell(row=target_row_idx, column=documents_col_idx)
        sheets_cell = worksheet.cell(row=target_row_idx, column=sheets_col_idx)

        number_cell.value = row_context.get("number", "")
        documents_cell.value = row_context.get("document_text", "")
        sheets_cell.value = row_context.get("sheets_count", "")

        self._set_cell_bold(cell=number_cell, bold=False)
        self._set_cell_bold(cell=sheets_cell, bold=False)
        self._set_cell_bold(
            cell=documents_cell,
            bold=bool(row_context.get("is_act_row")),
        )

    def _count_preallocated_table_rows(
        self,
        *,
        worksheet: Worksheet,
        template_row_idx: int,
        page_col_idx: int,
    ) -> int:
        """
        Считает, сколько строк табличной части уже предусмотрено в XLSX-шаблоне.

        В строках таблицы колонка "Страница" содержит формулу. Когда формула
        заканчивается, начинается нижний служебный блок реестра.
        """
        rows_count = 0
        current_row_idx = template_row_idx

        while current_row_idx <= worksheet.max_row:
            page_cell = worksheet.cell(row=current_row_idx, column=page_col_idx)
            if not isinstance(page_cell.value, str) or not page_cell.value.startswith("="):
                break

            rows_count += 1
            current_row_idx += 1

        if rows_count <= 0:
            raise RegistryXlsxRendererValidationError(
                "Не удалось определить диапазон табличной части реестра в шаблоне."
            )

        return rows_count

    def _set_cell_bold(self, *, cell, bold: bool) -> None:
        current_font = copy(cell.font) if cell.font else Font()
        current_font.bold = bold
        cell.font = current_font

    def _apply_documents_rows_heights(
        self,
        *,
        worksheet: Worksheet,
        template_row_idx: int,
        documents_col_idx: int,
        sheets_col_idx: int,
        rows_count: int,
    ) -> None:
        if rows_count <= 0:
            return

        workbook = worksheet.parent
        probe_title = "__registry_height_probe__"
        if probe_title in workbook.sheetnames:
            workbook.remove(workbook[probe_title])

        usable_width_px = max(
            36,
            int(
                self._calculate_usable_text_width_px(
                    worksheet=worksheet,
                    start_col_idx=documents_col_idx,
                    end_col_idx=sheets_col_idx - 1,
                )
                * 0.94
            ),
        )

        probe_sheet = workbook.create_sheet(title=probe_title)
        probe_sheet.sheet_state = "hidden"
        probe_sheet.column_dimensions["A"].width = self._pixels_to_excel_width(usable_width_px)

        base_height = (
            worksheet.row_dimensions[template_row_idx].height
            or worksheet.sheet_format.defaultRowHeight
            or 18
        )

        try:
            for row_offset in range(rows_count):
                source_row_idx = template_row_idx + row_offset
                source_cell = worksheet.cell(row=source_row_idx, column=documents_col_idx)
                probe_cell = probe_sheet.cell(row=row_offset + 1, column=1)

                probe_cell.value = source_cell.value
                probe_cell.font = copy(source_cell.font)
                probe_cell.alignment = copy(source_cell.alignment)

                measured_height = self._measure_probe_row_height(
                    probe_cell=probe_cell,
                    usable_width_px=usable_width_px,
                    is_act_row=bool(getattr(source_cell.font, "bold", False)),
                )
                probe_sheet.row_dimensions[row_offset + 1].height = measured_height
                worksheet.row_dimensions[source_row_idx].height = round(max(base_height, measured_height), 1)
        finally:
            workbook.remove(probe_sheet)

    def _measure_probe_row_height(
        self,
        *,
        probe_cell,
        usable_width_px: int,
        is_act_row: bool,
    ) -> float:
        pil_font = self._get_pil_font(
            font_name=getattr(probe_cell.font, "name", None),
            font_size=getattr(probe_cell.font, "sz", None),
            bold=bool(getattr(probe_cell.font, "bold", False)),
        )
        wrapped_lines = self._wrap_text_to_pixel_width(
            text=str(probe_cell.value or ""),
            max_width_px=usable_width_px,
            pil_font=pil_font,
        )
        ascent, descent = pil_font.getmetrics()
        line_height_multiplier = 1.33 if is_act_row else 1.28
        bottom_padding_px = 14 if is_act_row else 11
        line_height_px = max(1, int(math.ceil((ascent + descent) * line_height_multiplier)))
        total_height_px = (len(wrapped_lines) * line_height_px) + bottom_padding_px
        return round(self._pixels_to_points(total_height_px), 1)

    def _calculate_combined_excel_width(
        self,
        *,
        worksheet: Worksheet,
        start_col_idx: int,
        end_col_idx: int,
    ) -> float:
        total_width = 0.0
        for col_idx in range(start_col_idx, end_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            column_dimension = worksheet.column_dimensions[column_letter]
            total_width += column_dimension.width or worksheet.sheet_format.defaultColWidth or 8.43

        return total_width

    def _adjust_documents_row_height(
        self,
        *,
        worksheet: Worksheet,
        row_idx: int,
        template_row_idx: int,
        documents_col_idx: int,
        sheets_col_idx: int,
        text: str,
        is_bold: bool,
    ) -> None:
        """
        Для merged-ячейки B:D авто-высота работает нестабильно, поэтому
        оцениваем нужное число визуальных строк вручную и задаём height.
        """
        base_height = (
            worksheet.row_dimensions[template_row_idx].height
            or worksheet.sheet_format.defaultRowHeight
            or 18
        )
        estimated_lines_count = self._estimate_wrapped_lines_count(
            worksheet=worksheet,
            row_idx=row_idx,
            documents_col_idx=documents_col_idx,
            sheets_col_idx=sheets_col_idx,
            text=text,
            is_bold=is_bold,
        )

        worksheet.row_dimensions[row_idx].height = round(max(base_height, estimated_lines_count), 1)

    def _estimate_wrapped_lines_count(
        self,
        *,
        worksheet: Worksheet,
        row_idx: int,
        documents_col_idx: int,
        sheets_col_idx: int,
        text: str,
        is_bold: bool,
    ) -> float:
        cell_font = worksheet.cell(row=row_idx, column=documents_col_idx).font
        usable_width_px = max(
            36,
            self._calculate_usable_text_width_px(
                worksheet=worksheet,
                start_col_idx=documents_col_idx,
                end_col_idx=sheets_col_idx - 1,
            ),
        )
        pil_font = self._get_pil_font(
            font_name=getattr(cell_font, "name", None),
            font_size=getattr(cell_font, "sz", None),
            bold=is_bold or bool(getattr(cell_font, "bold", False)),
        )
        wrapped_lines = self._wrap_text_to_pixel_width(
            text=text,
            max_width_px=usable_width_px,
            pil_font=pil_font,
        )
        ascent, descent = pil_font.getmetrics()
        line_height_px = max(1, ascent + descent)
        total_height_px = (len(wrapped_lines) * line_height_px) + 4
        return self._pixels_to_points(total_height_px)

    def _calculate_usable_text_width_px(
        self,
        *,
        worksheet: Worksheet,
        start_col_idx: int,
        end_col_idx: int,
    ) -> int:
        total_width_px = 0
        for col_idx in range(start_col_idx, end_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            column_dimension = worksheet.column_dimensions[column_letter]
            excel_width = column_dimension.width or worksheet.sheet_format.defaultColWidth or 8.43
            total_width_px += self._excel_width_to_pixels(excel_width)

        return max(0, total_width_px - 10)

    def _wrap_text_to_pixel_width(
        self,
        *,
        text: str,
        max_width_px: int,
        pil_font,
    ) -> list[str]:
        normalized_text = (text or "").replace("\r\n", "\n").replace("\r", "\n")
        source_lines = normalized_text.split("\n") or [""]
        result: list[str] = []

        for source_line in source_lines:
            stripped_line = source_line.strip()
            if not stripped_line:
                result.append("")
                continue

            words = stripped_line.split()
            current_line = ""

            for word in words:
                candidate = word if not current_line else f"{current_line} {word}"
                if self._measure_text_width_px(candidate, pil_font) <= max_width_px:
                    current_line = candidate
                    continue

                if current_line:
                    result.append(current_line)

                if self._measure_text_width_px(word, pil_font) <= max_width_px:
                    current_line = word
                    continue

                split_word_lines = self._split_long_token(
                    token=word,
                    max_width_px=max_width_px,
                    pil_font=pil_font,
                )
                result.extend(split_word_lines[:-1])
                current_line = split_word_lines[-1]

            if current_line:
                result.append(current_line)

        return result or [""]

    def _split_long_token(
        self,
        *,
        token: str,
        max_width_px: int,
        pil_font,
    ) -> list[str]:
        chunks: list[str] = []
        current = ""

        for char in token:
            candidate = f"{current}{char}"
            if current and self._measure_text_width_px(candidate, pil_font) > max_width_px:
                chunks.append(current)
                current = char
                continue
            current = candidate

        if current:
            chunks.append(current)

        return chunks or [token]

    def _measure_text_width_px(self, text: str, pil_font) -> int:
        if not text:
            return 0
        return int(math.ceil(pil_font.getlength(text)))

    def _pixels_to_points(self, pixels: int | float) -> float:
        return float(pixels) * 72.0 / 96.0

    def _excel_width_to_pixels(self, width: float) -> int:
        if width <= 0:
            return 0
        return int(math.floor(width * 7 + 5))

    def _pixels_to_excel_width(self, pixels: int | float) -> float:
        if pixels <= 5:
            return 0.0
        return max(0.0, (float(pixels) - 5.0) / 7.0)

    @lru_cache(maxsize=32)
    def _resolve_font_file_path(self, font_name: str, bold: bool) -> str | None:
        normalized_name = (font_name or "").strip().lower()
        candidates = self.FONT_FILE_CANDIDATES.get(normalized_name, ())
        if not candidates:
            return None

        preferred_files = candidates[1::2] if bold else candidates[0::2]
        fallback_files = tuple(file_name for file_name in candidates if file_name not in preferred_files)

        for file_name in (*preferred_files, *fallback_files):
            for fonts_dir in self.FONT_SEARCH_DIRS:
                direct_path = fonts_dir / file_name
                if direct_path.exists():
                    return str(direct_path)

                if not fonts_dir.exists():
                    continue

                matches = list(fonts_dir.rglob(file_name))
                if matches:
                    return str(matches[0])

        return None

    @lru_cache(maxsize=64)
    def _load_pil_font(self, font_name: str, font_size: int, bold: bool):
        font_path = self._resolve_font_file_path(font_name, bold)
        if font_path:
            try:
                return ImageFont.truetype(font_path, font_size)
            except OSError:
                pass
        return ImageFont.load_default()

    def _get_pil_font(self, *, font_name: str | None, font_size: Any, bold: bool):
        normalized_size = 11
        if font_size:
            try:
                normalized_size = max(6, int(round(float(font_size))))
            except (TypeError, ValueError):
                normalized_size = 11

        normalized_name = (font_name or "Calibri").strip()
        return self._load_pil_font(normalized_name, normalized_size, bold)

    def _delete_rows_preserving_footer_layout(
        self,
        *,
        worksheet: Worksheet,
        delete_start_row: int,
        amount: int,
    ) -> None:
        """
        Удаляет хвост пустых строк таблицы, сохраняя нижний блок шаблона.

        Для footer-блока важно бережно обработать:
        - merged ranges ниже удаляемого диапазона;
        - индивидуальные высоты строк.

        Иначе openpyxl может сместить значения, но визуально "растянуть"
        объединённые ячейки и подписи внизу реестра.
        """
        if amount <= 0:
            return

        deleted_last_row = delete_start_row + amount - 1
        original_max_row = worksheet.max_row

        footer_merged_ranges = []
        for merged_range in list(worksheet.merged_cells.ranges):
            if merged_range.max_row < delete_start_row:
                continue

            if merged_range.min_row <= deleted_last_row:
                worksheet.unmerge_cells(str(merged_range))
                continue

            footer_merged_ranges.append(
                (
                    merged_range.min_row - amount,
                    merged_range.min_col,
                    merged_range.max_row - amount,
                    merged_range.max_col,
                )
            )
            worksheet.unmerge_cells(str(merged_range))

        footer_row_heights: dict[int, float] = {}
        for source_row_idx in range(delete_start_row + amount, original_max_row + 1):
            source_dimension = worksheet.row_dimensions[source_row_idx]
            if source_dimension.height is not None:
                footer_row_heights[source_row_idx - amount] = source_dimension.height

        worksheet.delete_rows(delete_start_row, amount=amount)

        for target_row_idx, height in footer_row_heights.items():
            worksheet.row_dimensions[target_row_idx].height = height

        for min_row, min_col, max_row, max_col in footer_merged_ranges:
            worksheet.merge_cells(
                start_row=min_row,
                start_column=min_col,
                end_row=max_row,
                end_column=max_col,
            )

    def _apply_bold_to_row(self, *, worksheet: Worksheet, row_idx: int) -> None:
        """
        Делает всю строку жирной.
        """
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell, MergedCell):
                continue

            current_font = copy(cell.font) if cell.font else Font()
            current_font.bold = True
            cell.font = current_font
