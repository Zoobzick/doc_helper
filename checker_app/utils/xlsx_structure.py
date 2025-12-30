from dataclasses import dataclass
from typing import IO, List, Tuple, Union

from openpyxl import load_workbook


@dataclass
class SheetItem:
    index: int        # (index) — порядковый номер листа
    title: str        # (title) — имя листа


@dataclass
class XlsxCellItem:
    sheet_index: int  # (sheet_index) — индекс листа
    sheet_title: str  # (sheet_title) — имя листа
    row: int          # (row) — номер строки (1..)
    col: int          # (col) — номер колонки (1..)
    coordinate: str   # (coordinate) — координата ячейки ("B3")
    value: str        # (value) — строковое представление значения/формулы


def _cell_to_text(cell) -> str:
    # (cell) — openpyxl.cell.cell.Cell
    if cell.value is None:
        return ""
    # формулы в openpyxl обычно как строка вида "=SUM(A1:A2)"
    return str(cell.value).strip()


def extract_xlsx_structure(source: Union[str, "bytes", IO]) -> Tuple[List[SheetItem], List[XlsxCellItem]]:
    """
    (source) — путь до файла или file-like объект (UploadedFile.file).
    Возвращаем:
      - (sheets) список листов
      - (cells) список непустых ячеек по всем листам
    """
    wb = load_workbook(filename=source, data_only=False, read_only=True)

    sheets: List[SheetItem] = []
    cells: List[XlsxCellItem] = []

    for si, ws in enumerate(wb.worksheets):
        sheets.append(SheetItem(index=si, title=ws.title))

        # (ws.iter_rows) — идём по используемому диапазону, read_only=True экономит память
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                text = _cell_to_text(cell)
                if text:  # пропускаем пустые
                    cells.append(
                        XlsxCellItem(
                            sheet_index=si,
                            sheet_title=ws.title,
                            row=cell.row,
                            col=cell.column,
                            coordinate=cell.coordinate,
                            value=text,
                        )
                    )

    wb.close()
    return sheets, cells
